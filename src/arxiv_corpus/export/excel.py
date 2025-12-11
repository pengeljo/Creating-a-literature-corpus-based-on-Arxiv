"""Excel export functionality."""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from arxiv_corpus.analysis.paragraph_search import ParagraphMatch
from arxiv_corpus.config import ExcelConfig
from arxiv_corpus.storage.models import Paper, SearchResult
from arxiv_corpus.utils.logging import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """Export data to Excel files."""

    def __init__(self, config: ExcelConfig | None = None) -> None:
        """Initialize Excel exporter.

        Args:
            config: Excel configuration.
        """
        self.config = config or ExcelConfig()

    def export_papers(
        self,
        papers: list[Paper],
        output_path: str | Path,
        include_abstract: bool = True,
    ) -> Path:
        """Export papers to Excel.

        Args:
            papers: List of papers to export.
            output_path: Output file path.
            include_abstract: Whether to include abstracts.

        Returns:
            Path to created file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Build DataFrame
        data: list[dict[str, Any]] = []
        for paper in papers:
            row = {
                "arxiv_id": paper.arxiv_id,
                "title": paper.title,
                "authors": ", ".join(a.name for a in paper.authors),
                "categories": ", ".join(paper.categories),
                "published_date": paper.published_date.isoformat() if paper.published_date else "",
                "status": paper.status.value,
                "occurrence_count": paper.occurrence_count,
                "search_queries": "; ".join(paper.search_queries),
                "pdf_url": paper.pdf_url or "",
            }
            if include_abstract:
                row["abstract"] = paper.abstract or ""
            data.append(row)

        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine="openpyxl")

        logger.info(f"Exported {len(papers)} papers to {output_path}")
        return output_path

    def export_search_results(
        self,
        results: list[SearchResult],
        output_path: str | Path,
    ) -> Path:
        """Export search results to Excel with cross-reference matrix.

        Args:
            results: List of search results.
            output_path: Output file path.

        Returns:
            Path to created file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # Collect all unique paper IDs
        all_paper_ids: set[str] = set()
        for result in results:
            all_paper_ids.update(result.paper_ids)

        # Build cross-reference matrix
        data: list[dict[str, Any]] = []
        for paper_id in sorted(all_paper_ids):
            row: dict[str, Any] = {"arxiv_id": paper_id}

            # Count occurrences
            count = sum(1 for r in results if paper_id in r.paper_ids)
            row["occurrence_count"] = count

            # Mark which queries found this paper
            for result in results:
                query_key = result.query[:50]  # Truncate for column name
                row[query_key] = "x" if paper_id in result.paper_ids else ""

            data.append(row)

        df = pd.DataFrame(data)

        # Sort by occurrence count descending
        df = df.sort_values("occurrence_count", ascending=False)

        df.to_excel(output_path, index=False, engine="openpyxl")

        logger.info(
            f"Exported cross-reference matrix with {len(all_paper_ids)} papers to {output_path}"
        )
        return output_path

    def export_paragraphs(
        self,
        paragraph_matches: list[ParagraphMatch],
        output_path: str | Path,
        include_text: bool = True,
    ) -> Path:
        """Export paragraph matches to Excel.

        Args:
            paragraph_matches: List of paragraph matches.
            output_path: Output file path.
            include_text: Whether to include full paragraph text.

        Returns:
            Path to created file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        data: list[dict[str, Any]] = []
        for pm in paragraph_matches:
            row = {
                "arxiv_id": pm.paragraph.arxiv_id,
                "paragraph_index": pm.paragraph.paragraph_index,
                "total_hits": pm.total_hits,
                "unique_terms": pm.unique_terms,
                "hit_summary": pm.hit_summary,
                "word_count": pm.paragraph.word_count,
                "sentence_count": pm.paragraph.sentence_count,
            }
            if include_text:
                row["text"] = pm.paragraph.text

            # Add individual term columns
            for match in pm.matches:
                col_name = f"{match.original_term}_r{match.rank}"
                row[col_name] = match.count

            data.append(row)

        df = pd.DataFrame(data)
        df.to_excel(output_path, index=False, engine="openpyxl")

        logger.info(f"Exported {len(paragraph_matches)} paragraph matches to {output_path}")
        return output_path

    def export_corpus_summary(
        self,
        papers: list[Paper],
        paragraph_count: int,
        hit_statistics: dict,
        output_path: str | Path,
    ) -> Path:
        """Export a summary workbook with multiple sheets.

        Args:
            papers: List of papers in corpus.
            paragraph_count: Total number of paragraphs.
            hit_statistics: Hit statistics dictionary.
            output_path: Output file path.

        Returns:
            Path to created file.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"  # type: ignore

        # Header styling
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

        summary_data = [
            ["Metric", "Value"],
            ["Total Papers", len(papers)],
            ["Total Paragraphs", paragraph_count],
            ["Paragraphs with Hits", hit_statistics.get("total_paragraphs", 0)],
            ["Total Term Hits", hit_statistics.get("total_hits", 0)],
            ["Unique Terms Matched", hit_statistics.get("unique_terms", 0)],
            ["Avg Hits per Paragraph", f"{hit_statistics.get('hits_per_paragraph', 0):.2f}"],
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)  # type: ignore
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        # Term frequencies sheet
        ws_terms = wb.create_sheet("Term Frequencies")
        ws_terms.append(["Term", "Frequency"])
        ws_terms.cell(row=1, column=1).font = header_font
        ws_terms.cell(row=1, column=2).font = header_font

        for term, freq in hit_statistics.get("term_frequencies", {}).items():
            ws_terms.append([term, freq])

        # Rank distribution sheet
        ws_ranks = wb.create_sheet("Rank Distribution")
        ws_ranks.append(["Rank", "Hit Count"])
        ws_ranks.cell(row=1, column=1).font = header_font
        ws_ranks.cell(row=1, column=2).font = header_font

        for rank, count in sorted(hit_statistics.get("rank_distribution", {}).items()):
            ws_ranks.append([rank, count])

        # Papers sheet (condensed)
        ws_papers = wb.create_sheet("Papers")
        paper_headers = ["arxiv_id", "title", "status", "occurrence_count"]
        ws_papers.append(paper_headers)
        for idx, _header in enumerate(paper_headers, 1):
            ws_papers.cell(row=1, column=idx).font = header_font
            ws_papers.cell(row=1, column=idx).fill = header_fill

        for paper in papers[:1000]:  # Limit to first 1000
            ws_papers.append(
                [
                    paper.arxiv_id,
                    paper.title[:100],  # Truncate long titles
                    paper.status.value,
                    paper.occurrence_count,
                ]
            )

        # Adjust column widths
        for ws in wb.worksheets:
            for column_cells in ws.columns:  # type: ignore
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 50)  # type: ignore

        wb.save(output_path)
        logger.info(f"Exported corpus summary to {output_path}")
        return output_path
